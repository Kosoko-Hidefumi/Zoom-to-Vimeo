"""
vimeo_results.csv から指定列を抜き出し Excel ファイルを生成する。

実行すると 2 ファイルを同時に出力する:
  vimeo_links.xlsx       … 全行
  vimeo_links_omit.xlsx  … link 列が空の行を除いた行のみ

使い方（zoom_download ディレクトリで）:
  python export_vimeo_excel.py
  python export_vimeo_excel.py --in-csv ../vimeo_results.csv --out-excel ../vimeo_links.xlsx
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit(
        "[ERROR] openpyxl が見つかりません。pip install openpyxl を実行してください。"
    )

EXTRACT_COLS = [
    "講師名（英語）",
    "専門科",
    "所属",
    "レクチャータイトル（英語）",
    "パスコード",
    "link",
]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT   = Font(color="0563C1", underline="single")
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")

COL_WIDTHS = {
    "講師名（英語）":       22,
    "専門科":              12,
    "所属":                40,
    "レクチャータイトル（英語）": 60,
    "パスコード":           12,
    "link":                38,
}


def write_excel(rows: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vimeo Links"

    for col_idx, col_name in enumerate(EXTRACT_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(EXTRACT_COLS, start=1):
            value = (row.get(col_name) or "").strip()
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_name in ("所属", "レクチャータイトル（英語）")),
            )
            if fill:
                cell.fill = fill
            if col_name == "link" and value.startswith("http"):
                cell.hyperlink = value
                cell.font = LINK_FONT

    for col_idx, col_name in enumerate(EXTRACT_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]

    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"保存しました: {out_path.resolve()}  （{len(rows)} 行）")


def parse_args():
    p = argparse.ArgumentParser(description="vimeo_results.csv → Excel")
    p.add_argument(
        "--in-csv",
        type=Path,
        default=Path("../vimeo_results.csv"),
        help="入力 CSV（既定: ../vimeo_results.csv）",
    )
    p.add_argument(
        "--out-excel",
        type=Path,
        default=Path("../vimeo_links.xlsx"),
        help="全行 Excel の出力先（既定: ../vimeo_links.xlsx）",
    )
    p.add_argument(
        "--uploaded-only",
        action="store_true",
        help="status が uploaded / skipped(already_uploaded) の行のみ対象にする",
    )
    return p.parse_args()


def main():
    args = parse_args()

    if not args.in_csv.is_file():
        raise SystemExit(f"[ERROR] CSV が見つかりません: {args.in_csv}")

    with open(args.in_csv, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        raise SystemExit("[ERROR] CSV にデータ行がありません。")

    missing = [c for c in EXTRACT_COLS if c not in rows[0]]
    if missing:
        raise SystemExit(f"[ERROR] CSV に列がありません: {missing}")

    if args.uploaded_only:
        rows = [
            r for r in rows
            if (r.get("status") or "").startswith("uploaded")
            or (r.get("status") or "").startswith("skipped(already_uploaded)")
        ]

    # 全行版
    write_excel(rows, args.out_excel)

    # link 空行を省いた版
    rows_with_link = [r for r in rows if (r.get("link") or "").strip().startswith("http")]
    omit_path = args.out_excel.with_stem(args.out_excel.stem + "_omit")
    write_excel(rows_with_link, omit_path)


if __name__ == "__main__":
    main()
